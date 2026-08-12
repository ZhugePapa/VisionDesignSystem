from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "Vision-Demo-全部功能路线图-v0.1.xlsx"
TODAY = date(2026, 8, 12)

BLUE = "2563EB"
BLUE_DARK = "1E3A8A"
BLUE_LIGHT = "DBEAFE"
NAVY = "172554"
GREEN = "16A34A"
GREEN_LIGHT = "DCFCE7"
AMBER = "D97706"
AMBER_LIGHT = "FEF3C7"
RED = "DC2626"
RED_LIGHT = "FEE2E2"
GREY_50 = "F8FAFC"
GREY_100 = "F1F5F9"
GREY_200 = "E2E8F0"
GREY_500 = "64748B"
GREY_700 = "334155"
WHITE = "FFFFFF"

FIGMA_BASE = "https://www.figma.com/design/AH5UD6NGdbtqy2z4o9k7QW/%E8%BD%AF%E4%BB%B6%E5%B7%A5%E5%8E%82"
DEMO_BASE = "http://127.0.0.1:5173/demo/#/projects/lighting"


HEADERS = [
    "功能编号", "一级模块", "二级模块", "功能名称", "类型", "核心演示路径", "实现边界",
    "使用场景与价值", "需求来源", "需求部门", "优先级", "状态", "当前成熟度", "目标成熟度",
    "当前实现说明", "下一步动作", "负责人", "协作角色", "计划版本", "排期状态", "目标完成时间",
    "设计稿", "Demo 入口", "验收标准", "依赖与风险", "最新进展／阻塞项", "验收结果",
    "实际完成时间", "最后更新时间",
]


def figma(node: str) -> str:
    return f"{FIGMA_BASE}?node-id={node.replace(':', '-')}"


def demo(path: str) -> str:
    return f"{DEMO_BASE}{path}"


rows: list[dict[str, object]] = []


def add(
    module: str,
    submodule: str,
    name: str,
    *,
    kind: str = "新功能",
    core: str = "否",
    boundary: str = "Demo 模拟",
    value: str,
    source: str = "跨部门推广规划",
    departments: str = "产品、研发",
    priority: str = "P2",
    status: str = "已规划",
    current: str = "L0 占位可达",
    target: str = "L3 数据联动",
    implementation: str = "当前导航可达，使用通用占位页面。",
    next_action: str = "补齐关键场景、示例数据和核心交互。",
    owner: str = "待认领",
    roles: str = "产品、设计、前端",
    version: str = "v0.4",
    schedule: str = "建议排期",
    target_date: date | None = date(2026, 10, 16),
    design: str = "",
    entry: str = "",
    acceptance: str = "可从导航进入；核心信息完整；关键交互可操作；数据在相关页面间保持一致。",
    risk: str = "需确认部门共识与设计稿；仅模拟核心流程，不建设真实后端。",
    progress: str = "待评审范围与负责人。",
    result: str = "未开始",
    actual_date: date | None = None,
) -> None:
    rows.append({
        "功能编号": f"DEMO-{len(rows) + 1:03d}",
        "一级模块": module,
        "二级模块": submodule,
        "功能名称": name,
        "类型": kind,
        "核心演示路径": core,
        "实现边界": boundary,
        "使用场景与价值": value,
        "需求来源": source,
        "需求部门": departments,
        "优先级": priority,
        "状态": status,
        "当前成熟度": current,
        "目标成熟度": target,
        "当前实现说明": implementation,
        "下一步动作": next_action,
        "负责人": owner,
        "协作角色": roles,
        "计划版本": version,
        "排期状态": schedule,
        "目标完成时间": target_date,
        "设计稿": design,
        "Demo 入口": entry,
        "验收标准": acceptance,
        "依赖与风险": risk,
        "最新进展／阻塞项": progress,
        "验收结果": result,
        "实际完成时间": actual_date,
        "最后更新时间": TODAY,
    })


def complete(
    module: str,
    submodule: str,
    name: str,
    *,
    value: str,
    entry: str,
    design: str = "",
    core: str = "是",
    boundary: str = "UI 交互 + 数据联动",
    implementation: str = "已完成页面、关键交互和示例数据联动。",
    acceptance: str = "页面可稳定访问；主要交互可操作；刷新和页面跳转后数据语义一致。",
    version: str = "v0.1",
    kind: str = "现有能力",
    departments: str = "产品、研发、设计",
) -> None:
    add(
        module, submodule, name,
        kind=kind, core=core, boundary=boundary, value=value, source="现有 Demo 实现",
        departments=departments, priority="P0" if core == "是" else "P1", status="已完成",
        current="L3 数据联动", target="L3 数据联动", implementation=implementation,
        next_action="纳入回归清单，持续根据部门反馈微调。", owner="Demo 小组",
        roles="产品、设计、前端", version=version, schedule="已确认", target_date=TODAY,
        design=design, entry=entry, acceptance=acceptance, risk="需防止后续修改破坏跨页面数据一致性。",
        progress="已完成并通过本地类型、构建与页面回归。", result="通过", actual_date=TODAY,
    )


# 基础体验与协作基础
complete("基础体验", "应用框架", "项目级应用壳层与导航", value="让不同部门在统一的信息架构中定位项目与研发活动。", entry=demo("/overview"))
complete("基础体验", "项目切换", "多项目切换与导航状态保持", value="演示同一套能力在不同项目上下文中的复用。", entry=demo("/overview"), core="否")
complete("基础体验", "主题", "浅色／深色主题切换", value="验证设计系统变量和组件在双主题下的一致性。", entry=demo("/overview"), core="否")
complete("基础体验", "路由", "关键页面深链接", value="会议和异步沟通时可直接分享具体功能页面。", entry=demo("/code/repositories/flight-control-core"))
add("基础体验", "全局搜索", "跨模块全局搜索", value="通过一个入口快速定位仓库、需求、流水线、应用与测试对象。", priority="P1", version="v0.3", target_date=date(2026, 9, 18), entry=demo("/overview"))
add("基础体验", "通知", "通知中心与待办提示", value="模拟评审、流水线失败、部署完成等跨模块提醒。", departments="产品、研发、测试、运维", priority="P2", version="v0.4")
add("基础体验", "响应式", "常用桌面分辨率适配", kind="体验优化", value="保证会议室大屏、笔记本和小屏窗口下都可顺畅演示。", departments="全体部门", priority="P1", current="L2 可交互", target="L3 数据联动", implementation="核心仓库页面已有小屏优化，但尚未形成统一回归矩阵。", version="v0.2", target_date=date(2026, 8, 28), entry=demo("/code/repositories/flight-control-core"))
add("基础体验", "无障碍", "键盘与可访问性基线", kind="体验优化", value="降低演示操作门槛，并保证组件语义与焦点状态一致。", priority="P2", version="v0.4")
complete("基础体验", "数据治理", "统一轻量示例数据源", value="让仓库、提交、分支、标签和合并请求呈现同一真实叙事。", entry=demo("/code/repositories/flight-control-core"), core="是", implementation="所有示例仓库已指向统一数据模型，计数和关联关系由数据计算。")
add("协作推广", "反馈闭环", "部门反馈收集与回链", kind="协作治理", value="让使用者从页面问题快速回链到路线图条目和设计稿。", departments="全体部门", priority="P0", status="已规划", current="L0 占位可达", target="L2 可交互", owner="Demo 运营", roles="产品、设计、研发代表", version="v0.2", target_date=date(2026, 8, 28), acceptance="提供统一反馈入口；每条反馈可关联功能编号、页面、截图、提出部门与处理状态。")
add("协作推广", "变更沟通", "版本更新日志", kind="协作治理", value="让部门清楚每次更新新增了什么、调整了什么、是否影响既有演示路径。", departments="全体部门", priority="P1", owner="Demo 运营", version="v0.2", target_date=date(2026, 8, 28), target="L2 可交互")
add("协作推广", "使用支持", "Demo 使用指南与讲解脚本", kind="内容补充", value="帮助非建设人员独立完成核心场景演示。", departments="全体部门", priority="P1", owner="产品", roles="产品、设计、业务代表", version="v0.2", target_date=date(2026, 8, 28), target="L2 可交互")

# 项目与需求
complete("项目协作", "项目概览", "项目概览仪表盘", value="用需求、流水线、缺陷和环境指标建立项目全局认知。", entry=demo("/overview"), core="否", boundary="静态展示 + 局部交互")
add("项目协作", "项目管理", "项目管理总览", value="集中展示项目成员、迭代、里程碑和近期活动。", departments="产品、项目管理、研发", priority="P2", entry=demo("/management"))
add("项目协作", "项目管理", "成员与角色", value="模拟项目成员、角色和职责分工，支撑跨部门协作讨论。", departments="项目管理、研发", priority="P3", version="v1.0", target_date=date(2026, 10, 30))
add("项目协作", "项目管理", "迭代与里程碑", value="用统一时间轴连接需求、研发、测试和发布计划。", departments="产品、项目管理、研发、测试", priority="P1", version="v0.4")
add("需求管理", "需求列表", "需求列表与筛选", value="按状态、负责人、优先级和迭代浏览需求。", departments="产品、研发、测试", priority="P1", entry=demo("/requirements"), version="v0.3", target_date=date(2026, 9, 18))
add("需求管理", "需求详情", "需求详情与状态流转", value="围绕一个需求展示描述、负责人、验收标准和协作记录。", departments="产品、研发、测试", priority="P1", version="v0.3", target_date=date(2026, 9, 18))
add("需求管理", "追溯关系", "需求—代码—测试—发布追溯", value="展示需求如何关联提交、合并请求、测试结果和发布版本。", departments="产品、研发、测试、质量", priority="P0", core="是", target="L4 端到端叙事", version="v1.0", target_date=date(2026, 10, 30), acceptance="从需求详情可进入关联提交、合并请求、测试结果和发布记录；对象编号与状态一致。")
add("项目协作", "项目设置", "项目属性、通知与权限展示", value="支持讨论项目级配置和治理策略。", departments="项目管理、研发管理", priority="P3", entry=demo("/settings"), version="v1.0", target_date=date(2026, 10, 30))

# 代码仓库
complete("代码仓库", "仓库列表", "仓库目录与层级列表", value="浏览组织、分组和仓库层级，快速进入目标仓库。", entry=demo("/code/repositories"))
complete("代码仓库", "仓库列表", "仓库关键字搜索与排序", value="在多仓库场景下快速定位目标。", entry=demo("/code/repositories"), core="否")
add("代码仓库", "仓库列表", "高级筛选与收藏视图", value="按语言、活跃度、负责人、收藏状态筛选仓库。", priority="P2", current="L1 静态展示", target="L2 可交互", implementation="已预留筛选和设置入口，部分按钮仍为禁用演示态。", version="v0.3", target_date=date(2026, 9, 18), entry=demo("/code/repositories"))
add("代码仓库", "仓库创建", "新建／导入仓库弹窗", value="完整演示从代码仓库入口创建或导入仓库的流程。", priority="P2", target="L2 可交互", version="v0.3", target_date=date(2026, 9, 18), entry=demo("/code/repositories"), acceptance="点击新建仓库可打开表单；确认后新增一条本地示例仓库并给出成功反馈；无需真实 Git 服务。")
complete("代码仓库", "仓库浏览", "文件树、目录与文件内容浏览", value="模拟真实仓库的日常代码阅读路径。", entry=demo("/code/repositories/flight-control-core"))
complete("代码仓库", "仓库浏览", "分支／标签切换下拉框", value="切换引用后文件、提交和计数同步变化。", entry=demo("/code/repositories/flight-control-core"), design=figma("524:35227"))
complete("代码仓库", "仓库浏览", "代码克隆下拉框", value="展示 SSH／HTTP 克隆地址并支持复制。", entry=demo("/code/repositories/flight-control-core"), design=figma("524:35243"), core="否")
complete("代码仓库", "文件详情", "打开文件—代码内容", value="查看文件名称、大小、内容和操作区。", entry=demo("/code/repositories/flight-control-core"))
complete("代码仓库", "文件详情", "打开文件—修改追溯", value="按提交与作者展示逐行归属，模拟 blame 场景。", entry=demo("/code/repositories/flight-control-core"), design=figma("475:4006"))
complete("代码仓库", "文件详情", "打开文件—文件历史", value="以提交项展示当前文件的修改历史。", entry=demo("/code/repositories/flight-control-core"), design=figma("524:35732"))
add("代码仓库", "文件详情", "文件差异对比", value="从文件历史或提交详情进入新旧版本 diff，辅助评审沟通。", priority="P0", core="是", current="L0 占位可达", target="L3 数据联动", version="v0.2", target_date=date(2026, 8, 28), acceptance="可选择两个版本；展示增删行语义色、文件路径和提交信息；复用 Code Experience 组件。")
add("代码仓库", "文件详情", "仓库内代码搜索", value="按文件名或内容快速定位代码片段。", priority="P2", version="v0.3", target_date=date(2026, 9, 18))
add("代码仓库", "文件详情", "编辑、上传与下载演示", value="展示常见文件操作入口及成功反馈，但不写入真实仓库。", priority="P3", status="待评估", target="L2 可交互", version="v1.0", target_date=date(2026, 10, 30), risk="需要明确哪些写操作值得模拟，避免让使用者误认为已接入真实 Git。")
complete("代码仓库", "提交", "提交列表、筛选与描述展开", value="按分支和日期查看提交，展开描述并复制提交号。", entry=demo("/code/repositories/flight-control-core/commits?branch=main"), design=figma("486:5192"))
add("代码仓库", "提交", "提交详情与变更文件", value="查看单次提交说明、关联请求、文件列表和 diff。", priority="P0", core="是", version="v0.2", target_date=date(2026, 8, 28), design=figma("486:5192"), acceptance="点击提交标题进入详情；变更文件、增删统计、关联合并请求与文件 diff 数据一致。")
complete("代码仓库", "分支", "分支列表与真实计数", value="查看分支状态、更新人、合并请求和领先／滞后。", entry=demo("/code/repositories/flight-control-core/branches"), design=figma("494:33747"))
complete("代码仓库", "分支", "新建分支弹窗", value="模拟从指定来源创建分支并反馈成功。", entry=demo("/code/repositories/flight-control-core/branches"), design=figma("524:35535"), core="否", boundary="UI 交互")
complete("代码仓库", "分支", "保护分支与删除禁用态", value="表达默认分支、保护分支的操作限制。", entry=demo("/code/repositories/flight-control-core/branches"), core="否")
complete("代码仓库", "标签", "标签列表与描述展开", value="查看标签、创建人、提交、校验状态和下载入口。", entry=demo("/code/repositories/flight-control-core/tags"), design=figma("519:5677"))
complete("代码仓库", "标签", "新建标签弹窗", value="模拟从分支或提交创建标签并反馈成功。", entry=demo("/code/repositories/flight-control-core/tags"), design=figma("524:35535"), core="否", boundary="UI 交互")
complete("代码仓库", "合并请求", "合并请求列表与状态筛选", value="查看开启、已合并、已关闭等请求，并呈现分支流和变更量。", entry=demo("/code/repositories/flight-control-core/merge-requests"), design=figma("524:7424"))
add("代码仓库", "合并请求", "合并请求详情", value="围绕一个请求展示描述、提交、变更、评审和讨论。", priority="P0", core="是", target="L4 端到端叙事", version="v0.2", target_date=date(2026, 8, 28), design=figma("161:4885"), acceptance="列表标题可进入详情；详情包含概要、提交、文件变更、评审状态和讨论；关联对象可跳转。")
add("代码仓库", "合并请求", "新建合并请求流程", value="模拟选择源分支、目标分支、评审人并创建请求。", priority="P1", target="L3 数据联动", version="v0.3", target_date=date(2026, 9, 18), acceptance="确认后本地新增请求，页头计数、分支列表关联请求和请求列表同步更新。")
add("代码仓库", "评审记录", "代码评审记录", value="展示评审意见、处理状态和变更位置。", priority="P1", target="L3 数据联动", version="v0.3", target_date=date(2026, 9, 18))
add("代码仓库", "关联工作项", "代码与工作项关联", value="从仓库或合并请求追溯到需求、任务与缺陷。", departments="产品、研发、测试", priority="P0", core="是", target="L4 端到端叙事", version="v1.0", target_date=date(2026, 10, 30))
add("代码仓库", "入库记录", "代码入库记录", value="展示版本基线、审批状态和入库结果。", departments="研发、质量、配置管理", priority="P2", version="v0.4")
add("代码仓库", "统计", "仓库活跃度与贡献统计", value="展示提交趋势、贡献者、分支和合并请求统计。", departments="研发、研发管理", priority="P2", version="v0.4")
complete("代码仓库", "设置", "Webhook 设置与触发记录", value="演示服务集成配置和触发历史。", entry=demo("/code/repositories/flight-control-core?tab=settings"), core="否")
add("代码仓库", "设置", "仓库基本设置与通知设置", value="展示仓库属性、默认分支、通知和访问策略。", priority="P3", version="v1.0", target_date=date(2026, 10, 30), entry=demo("/code/repositories/flight-control-core?tab=settings"))

# 代码质量
add("代码质量", "质量总览", "代码质量仪表盘", value="聚合质量门禁、问题、覆盖率和趋势，建立研发质量共识。", departments="研发、测试、质量", priority="P1", core="是", version="v0.3", target_date=date(2026, 9, 18), entry=demo("/code/quality"))
add("代码质量", "问题管理", "静态检查问题列表与详情", value="按严重级别、规则、文件和负责人查看问题。", departments="研发、质量", priority="P1", version="v0.3", target_date=date(2026, 9, 18))
add("代码质量", "质量门禁", "质量门禁结果", value="解释一次提交或合并请求为何通过／未通过。", departments="研发、质量", priority="P1", version="v0.3", target_date=date(2026, 9, 18))
add("代码质量", "覆盖率", "覆盖率与质量趋势", value="按分支和时间展示覆盖率、缺陷密度和技术债变化。", departments="研发、测试、质量", priority="P2", version="v0.4")

# 流水线
add("流水线", "流水线列表", "流水线列表与筛选", value="查看流水线状态、触发方式、分支和最近运行。", departments="研发、测试、运维", priority="P0", core="是", version="v0.3", target_date=date(2026, 9, 18), entry=demo("/pipelines"))
add("流水线", "运行详情", "阶段图与运行状态", value="用阶段视图解释构建、测试、扫描和发布过程。", departments="研发、测试、运维", priority="P0", core="是", target="L4 端到端叙事", version="v0.3", target_date=date(2026, 9, 18))
add("流水线", "运行详情", "任务日志与失败定位", value="展开任务日志并定位失败步骤，支撑问题沟通。", departments="研发、测试、运维", priority="P1", version="v0.3", target_date=date(2026, 9, 18))
add("流水线", "运行操作", "手动执行、停止与重试", value="模拟常见流水线操作和状态反馈。", departments="研发、测试、运维", priority="P2", boundary="UI 交互 + 本地状态", target="L2 可交互", version="v0.4")
add("流水线", "流水线配置", "可视化编排与 YAML 预览", value="帮助讨论流水线阶段、任务和配置结构。", departments="研发、运维", priority="P3", status="待评估", version="v1.0", target_date=date(2026, 10, 30))

# 制品与依赖
add("制品管理", "制品仓库", "制品仓库与版本列表", value="浏览构建产物、版本、来源流水线和发布时间。", departments="研发、测试、运维、配置管理", priority="P1", core="是", version="v0.4", entry=demo("/artifacts/repository"))
add("制品管理", "制品详情", "版本详情、校验与下载", value="展示制品元数据、校验值、关联提交和下载入口。", departments="研发、测试、运维、配置管理", priority="P1", version="v0.4")
add("制品管理", "依赖库", "项目依赖与版本清单", value="查看项目直接／间接依赖、版本和使用范围。", departments="研发、质量、安全", priority="P2", version="v0.4", entry=demo("/artifacts/dependencies"))
add("制品管理", "依赖治理", "SBOM 与许可证", value="展示软件物料清单、许可证类型和合规风险。", departments="研发、质量、安全、法务", priority="P2", version="v1.0", target_date=date(2026, 10, 30))
add("制品管理", "依赖治理", "漏洞与修复建议", value="关联依赖漏洞、影响范围和建议升级版本。", departments="研发、质量、安全", priority="P2", version="v1.0", target_date=date(2026, 10, 30))

# 部署
complete("部署管理", "应用管理", "应用列表与部署状态", value="查看应用、环境、版本、部署进度和操作菜单。", entry=demo("/deployments/applications"), core="是")
complete("部署管理", "应用详情", "应用概览", value="展示应用运行概况、版本和环境信息。", entry=demo("/deployments/applications/flight-control/overview"), core="是")
complete("部署管理", "环境规划", "应用环境规划", value="展示开发、测试、生产环境及部署状态。", entry=demo("/deployments/applications/flight-control/environment-planning"), core="是")
add("部署管理", "发布编排", "发布阶段与审批编排", value="串联制品选择、环境、审批、部署和验证。", departments="研发、测试、运维、质量", priority="P0", core="是", target="L4 端到端叙事", version="v0.4", target_date=date(2026, 10, 16))
add("部署管理", "环境参数", "环境参数与差异对比", value="展示不同环境配置及差异，降低发布沟通成本。", departments="研发、测试、运维", priority="P1", version="v0.4")
add("部署管理", "发布历史", "发布历史与结果详情", value="按应用和环境追踪发布版本、发起人、耗时和结果。", departments="研发、测试、运维", priority="P1", version="v0.4")
add("部署管理", "环境管理", "环境列表与资源概况", value="集中查看环境类型、资源、状态和所属应用。", departments="测试、运维", priority="P2", version="v0.4", entry=demo("/deployments/environments"))
add("部署管理", "部署操作", "部署日志、验证与回滚", value="展示失败定位、发布验证和回滚路径。", departments="研发、测试、运维", priority="P0", core="是", target="L4 端到端叙事", version="v1.0", target_date=date(2026, 10, 30))

# 测试
add("测试管理", "测试计划", "测试计划列表与详情", value="按版本组织测试范围、负责人、周期和进度。", departments="测试、产品、研发", priority="P1", core="是", version="v0.4", entry=demo("/testing/plans"))
add("测试管理", "测试用例", "测试用例库与步骤详情", value="展示前置条件、步骤、预期结果和关联需求。", departments="测试、产品、研发", priority="P1", version="v0.4", entry=demo("/testing/cases"))
add("测试管理", "测试执行", "执行记录与结果", value="记录测试执行、通过率、失败原因和证据。", departments="测试、质量", priority="P1", version="v0.4")
add("测试管理", "测试缺陷", "缺陷列表、详情与回归状态", value="跟踪缺陷优先级、修复版本和回归结果。", departments="测试、研发、产品", priority="P1", version="v0.4", entry=demo("/testing/defects"))
add("测试管理", "测试报告", "版本测试报告", value="汇总计划、执行、缺陷和风险，支持发布决策。", departments="测试、质量、项目管理", priority="P2", version="v1.0", target_date=date(2026, 10, 30))

# 跨部门推广
add("协作推广", "路线图", "全部功能路线图", kind="协作治理", value="统一追踪现有能力、未来计划、状态、负责人、依赖和验收结果。", departments="全体部门", priority="P0", status="已完成", current="L2 可交互", target="L2 可交互", implementation="已生成 Excel 初版，包含主表、仪表盘、里程碑、字段说明和枚举字典。", next_action="组织首次评审，确认负责人、优先级和计划版本。", owner="Demo 运营", roles="各部门代表", version="v0.1", schedule="已确认", target_date=TODAY, acceptance="路线图覆盖现有导航与规划模块；字段可筛选；状态和优先级有统一枚举；每条功能有验收标准。", risk="负责人和排期仍需部门评审确认。", progress="初版已生成。", result="通过", actual_date=TODAY)
add("协作推广", "评审机制", "双周路线图评审", kind="协作治理", value="定期确认新增诉求、优先级、阻塞和验收结果。", departments="全体部门", priority="P0", boundary="协作机制", target="L2 可交互", owner="Demo 运营", roles="产品、设计、研发、测试、运维、业务代表", version="v0.2", target_date=date(2026, 8, 28), acceptance="形成固定会议节奏；每次更新路线图状态、负责人、排期与结论；超期项有原因。")
add("协作推广", "试点", "跨部门试点与反馈复盘", kind="协作治理", value="验证 Demo 是否能支撑需求讨论、设计评审和研发流程沟通。", departments="全体部门", priority="P0", core="是", boundary="协作机制", target="L4 端到端叙事", owner="Demo 运营", roles="各部门代表", version="v1.0", target_date=date(2026, 10, 30), acceptance="至少完成 3 个部门试点；收集任务完成率、关键问题和满意度；形成下一版改进清单。")
add("协作推广", "数据标准", "示例数据命名与一致性规范", kind="数据完善", value="避免新增页面出现时间、人员、编号和对象关系互相矛盾。", departments="产品、设计、研发", priority="P0", boundary="数据规范", target="L3 数据联动", owner="Demo 小组", version="v0.2", target_date=date(2026, 8, 28), acceptance="定义人员、项目、仓库、需求、流水线、制品、部署和测试对象的唯一编号与关联规则。")


ENUMS = {
    "状态": ["待评估", "已规划", "设计中", "开发中", "待验收", "已完成", "已暂停", "不采纳"],
    "优先级": ["P0", "P1", "P2", "P3"],
    "类型": ["现有能力", "新功能", "体验优化", "数据完善", "内容补充", "Bug", "协作治理"],
    "核心演示路径": ["是", "否"],
    "实现边界": ["静态展示", "UI 交互", "UI 交互 + 本地状态", "UI 交互 + 数据联动", "Demo 模拟", "协作机制", "数据规范"],
    "成熟度": ["L0 占位可达", "L1 静态展示", "L2 可交互", "L3 数据联动", "L4 端到端叙事"],
    "排期状态": ["已确认", "建议排期", "待排期", "不排期"],
    "验收结果": ["未开始", "待验收", "通过", "需修改", "不适用"],
}


FIELD_GUIDE = [
    ("功能编号", "自动编号", "稳定引用功能条目，会议和反馈中优先使用编号。", "必填"),
    ("一级模块", "单选", "项目级能力域，如代码仓库、流水线、部署管理。", "必填"),
    ("二级模块", "单选／文本", "一级模块下的功能分区。", "必填"),
    ("功能名称", "单行文本", "使用动宾结构，描述明确可验收的能力。", "必填"),
    ("类型", "单选", "区分现有能力、新功能、体验优化、数据完善等。", "必填"),
    ("核心演示路径", "是／否", "是否属于跨部门演示必须打通的主路径。", "必填"),
    ("实现边界", "单选", "明确仅做静态、交互、数据联动或协作机制，防止演变成真实后台项目。", "必填"),
    ("使用场景与价值", "多行文本", "解释该功能帮助谁完成什么沟通任务。", "必填"),
    ("需求来源", "文本", "记录设计稿、现有实现、部门诉求或路线图规划。", "建议"),
    ("需求部门", "多选文本", "标识关注和受益部门，用中文顿号分隔。", "建议"),
    ("优先级", "单选", "P0 主路径阻塞；P1 高价值；P2 增强；P3 可延后。", "必填"),
    ("状态", "单选", "统一生命周期：待评估→已规划→设计中→开发中→待验收→已完成。", "必填"),
    ("当前成熟度", "单选", "L0 占位；L1 静态；L2 交互；L3 数据联动；L4 端到端叙事。", "必填"),
    ("目标成熟度", "单选", "本次路线图承诺达到的演示成熟度。", "必填"),
    ("当前实现说明", "多行文本", "客观描述现在能演示到什么程度。", "建议"),
    ("下一步动作", "多行文本", "最近一个明确可执行动作。", "建议"),
    ("负责人", "人员", "对结果负责；未明确时保留“待认领”。", "必填"),
    ("协作角色", "多选文本", "设计、前端、产品、测试、运维、业务代表等。", "建议"),
    ("计划版本", "单选／文本", "所属演示里程碑。", "必填"),
    ("排期状态", "单选", "区分已确认日期和建议日期，避免把初版路线图误认为承诺。", "必填"),
    ("目标完成时间", "日期", "建议节奏需在评审后转为已确认。", "建议"),
    ("设计稿", "URL", "链接到具体 Figma 节点，不只链接文件首页。", "建议"),
    ("Demo 入口", "URL", "链接到可直接验证的页面。", "建议"),
    ("验收标准", "多行文本", "描述完成的可观察条件，避免仅写“完成开发”。", "必填"),
    ("依赖与风险", "多行文本", "记录设计、数据、组件、范围和跨部门决策依赖。", "建议"),
    ("最新进展／阻塞项", "多行文本", "每次评审更新，明确是否需要外部决策。", "建议"),
    ("验收结果", "单选", "未开始、待验收、通过、需修改、不适用。", "必填"),
    ("实际完成时间", "日期", "用于统计交付节奏。", "完成时必填"),
    ("最后更新时间", "自动日期", "判断信息是否陈旧。", "必填"),
]


MILESTONES = [
    ("v0.1", "仓库场景基线", date(2026, 8, 12), "已完成", "项目壳层、统一仓库数据、文件浏览、追溯、历史、提交、分支、标签、合并请求列表、应用管理基础。", "形成可用于内部评审的代码仓库主路径。"),
    ("v0.2", "仓库评审闭环", date(2026, 8, 28), "建议排期", "提交详情、文件 diff、合并请求详情、协作反馈、使用指南、数据标准。", "仓库从浏览扩展到变更评审闭环。"),
    ("v0.3", "研发协同链路", date(2026, 9, 18), "建议排期", "需求列表与详情、代码质量、流水线列表与运行详情、新建合并请求。", "打通需求—代码—构建的核心叙事。"),
    ("v0.4", "交付与验证链路", date(2026, 10, 16), "建议排期", "制品、部署编排、环境、测试计划、用例、缺陷与发布历史。", "打通构建—制品—部署—测试。"),
    ("v1.0", "跨部门试点版", date(2026, 10, 30), "建议排期", "端到端追溯、关键统计、治理能力、试点反馈闭环。", "支持多个部门围绕同一 Demo 独立沟通协作。"),
]


def style_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=10, color=GREY_700)
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE_LIGHT)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=WHITE))


def build_workbook() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 主表
    ws = wb.create_sheet("全部功能路线图")
    style_title(
        ws,
        "Vision Demo 全部功能路线图 v0.1",
        "初版依据当前代码、导航和已实现页面整理。未来日期均为建议节奏，只有“排期状态=已确认”才视为正式承诺；实现边界默认面向 Demo 模拟，不代表建设真实业务后端。",
        len(HEADERS),
    )
    for col, header in enumerate(HEADERS, 1):
        ws.cell(4, col, header)
    style_header(ws[4])
    ws.row_dimensions[4].height = 34

    for row_index, item in enumerate(rows, 5):
        for col_index, header in enumerate(HEADERS, 1):
            value = item.get(header, "")
            cell = ws.cell(row_index, col_index, value)
            cell.font = Font(name="Arial", size=9, color=GREY_700)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color=GREY_200))
            if isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
        ws.row_dimensions[row_index].height = 52
        for col_index in (1, 6, 11, 12, 13, 14, 19, 20, 27):
            ws.cell(row_index, col_index).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_index in (22, 23):
            cell = ws.cell(row_index, col_index)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    ws.freeze_panes = "E5"
    ws.auto_filter.ref = f"A4:AC{4 + len(rows)}"
    widths = {
        "A": 12, "B": 14, "C": 16, "D": 28, "E": 13, "F": 12, "G": 20, "H": 38, "I": 18,
        "J": 22, "K": 10, "L": 12, "M": 16, "N": 16, "O": 40, "P": 38, "Q": 14, "R": 22,
        "S": 12, "T": 12, "U": 14, "V": 32, "W": 36, "X": 44, "Y": 38, "Z": 36, "AA": 12,
        "AB": 14, "AC": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    table = Table(displayName="DemoRoadmap", ref=f"A4:AC{4 + len(rows)}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)

    # 数据验证字典
    enum_ws = wb.create_sheet("枚举字典")
    max_enum_len = max(len(values) for values in ENUMS.values())
    for col_index, (name, values) in enumerate(ENUMS.items(), 1):
        enum_ws.cell(1, col_index, name)
        enum_ws.cell(1, col_index).font = Font(bold=True, color=WHITE)
        enum_ws.cell(1, col_index).fill = PatternFill("solid", fgColor=BLUE_DARK)
        for row_index, value in enumerate(values, 2):
            enum_ws.cell(row_index, col_index, value)
        enum_ws.column_dimensions[enum_ws.cell(1, col_index).column_letter].width = 24
    enum_ws.freeze_panes = "A2"

    validation_map = {
        "E": ("类型", 5), "F": ("核心演示路径", 6), "G": ("实现边界", 7), "K": ("优先级", 11),
        "L": ("状态", 12), "M": ("成熟度", 13), "N": ("成熟度", 14), "T": ("排期状态", 20),
        "AA": ("验收结果", 27),
    }
    enum_col_by_name = {name: index for index, name in enumerate(ENUMS, 1)}
    for column_letter, (enum_name, _) in validation_map.items():
        enum_col = enum_ws.cell(1, enum_col_by_name[enum_name]).column_letter
        enum_end = 1 + len(ENUMS[enum_name])
        dv = DataValidation(type="list", formula1=f"'枚举字典'!${enum_col}$2:${enum_col}${enum_end}", allow_blank=False)
        dv.error = "请选择枚举字典中的值"
        dv.errorTitle = "无效选项"
        ws.add_data_validation(dv)
        dv.add(f"{column_letter}5:{column_letter}500")

    # 条件格式
    last_row = 4 + len(rows)
    status_colors = {
        "已完成": (GREEN_LIGHT, GREEN), "待验收": (BLUE_LIGHT, BLUE), "开发中": (BLUE_LIGHT, BLUE),
        "设计中": (AMBER_LIGHT, AMBER), "已规划": (GREY_100, GREY_700), "待评估": (GREY_100, GREY_500),
        "已暂停": (RED_LIGHT, RED), "不采纳": (GREY_200, GREY_500),
    }
    for status, (fill, font_color) in status_colors.items():
        ws.conditional_formatting.add(
            f"L5:L{last_row}",
            FormulaRule(formula=[f'$L5="{status}"'], fill=PatternFill("solid", fgColor=fill), font=Font(color=font_color, bold=True)),
        )
    priority_colors = {"P0": RED_LIGHT, "P1": AMBER_LIGHT, "P2": BLUE_LIGHT, "P3": GREY_100}
    for priority, fill in priority_colors.items():
        ws.conditional_formatting.add(
            f"K5:K{last_row}",
            FormulaRule(formula=[f'$K5="{priority}"'], fill=PatternFill("solid", fgColor=fill), font=Font(bold=True)),
        )

    # 仪表盘
    dash = wb.create_sheet("路线图仪表盘", 0)
    style_title(dash, "Vision Demo 路线图仪表盘", "数据自动汇总自“全部功能路线图”。请在主表维护状态、优先级、计划版本和负责人。", 12)
    dash.sheet_view.showGridLines = False
    dash.column_dimensions["A"].width = 3
    for col in range(2, 13):
        dash.column_dimensions[get_column_letter(col)].width = 16

    cards = [
        ("B4", "功能总数", f'=COUNTA(\'全部功能路线图\'!$A$5:$A${last_row})', NAVY),
        ("E4", "已完成", f'=COUNTIF(\'全部功能路线图\'!$L$5:$L${last_row},"已完成")', GREEN),
        ("H4", "P0／P1", f'=COUNTIF(\'全部功能路线图\'!$K$5:$K${last_row},"P0")+COUNTIF(\'全部功能路线图\'!$K$5:$K${last_row},"P1")', AMBER),
        ("K4", "待认领", f'=COUNTIF(\'全部功能路线图\'!$Q$5:$Q${last_row},"待认领")', RED),
    ]
    for anchor, label, formula, color in cards:
        cell = dash[anchor]
        col = cell.column
        row = cell.row
        dash.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        dash.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        dash.cell(row, col, label)
        dash.cell(row, col).fill = PatternFill("solid", fgColor=color)
        dash.cell(row, col).font = Font(color=WHITE, bold=True, size=10)
        dash.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        dash.cell(row + 1, col, formula)
        dash.cell(row + 1, col).fill = PatternFill("solid", fgColor=GREY_50)
        dash.cell(row + 1, col).font = Font(color=color, bold=True, size=22)
        dash.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")

    dash["B9"] = "状态分布"
    dash["B9"].font = Font(size=12, bold=True, color=NAVY)
    for i, status in enumerate(ENUMS["状态"], 10):
        dash.cell(i, 2, status)
        dash.cell(i, 3, f'=COUNTIF(\'全部功能路线图\'!$L$5:$L${last_row},B{i})')
    style_header(dash[9][1:3])

    modules = list(dict.fromkeys(str(row["一级模块"]) for row in rows))
    dash["E9"] = "模块分布"
    dash["E9"].font = Font(size=12, bold=True, color=NAVY)
    for i, module in enumerate(modules, 10):
        dash.cell(i, 5, module)
        dash.cell(i, 6, f'=COUNTIF(\'全部功能路线图\'!$B$5:$B${last_row},E{i})')
    style_header(dash[9][4:6])

    doughnut = DoughnutChart()
    doughnut.title = "状态分布"
    doughnut.holeSize = 55
    doughnut.add_data(Reference(dash, min_col=3, min_row=9, max_row=9 + len(ENUMS["状态"])), titles_from_data=True)
    doughnut.set_categories(Reference(dash, min_col=2, min_row=10, max_row=9 + len(ENUMS["状态"])))
    doughnut.height = 7
    doughnut.width = 10
    dash.add_chart(doughnut, "H9")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "各模块功能数量"
    bar.add_data(Reference(dash, min_col=6, min_row=9, max_row=9 + len(modules)), titles_from_data=True)
    bar.set_categories(Reference(dash, min_col=5, min_row=10, max_row=9 + len(modules)))
    bar.height = 7
    bar.width = 10
    dash.add_chart(bar, "H24")

    dash["B22"] = "首轮评审建议"
    dash["B22"].font = Font(size=12, bold=True, color=NAVY)
    recommendations = [
        "1. 先确认 P0 主路径是否覆盖部门沟通需要。",
        "2. 将“待认领”条目分配到角色或具体负责人。",
        "3. 将确认后的日期从“建议排期”改为“已确认”。",
        "4. 对每个 v0.2 条目补齐设计稿和验收人。",
        "5. 双周复盘一次状态、阻塞与范围边界。",
    ]
    for i, text in enumerate(recommendations, 23):
        dash.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
        dash.cell(i, 2, text)
        dash.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")
        dash.cell(i, 2).font = Font(size=10, color=GREY_700)
        dash.cell(i, 2).fill = PatternFill("solid", fgColor=GREY_50 if i % 2 else GREY_100)

    # 里程碑
    milestones = wb.create_sheet("里程碑")
    style_title(milestones, "Vision Demo 建议里程碑", "v0.2 及以后日期为初版建议，需要在路线图评审中确认。", 6)
    milestone_headers = ["版本", "主题", "建议完成时间", "排期状态", "主要范围", "里程碑验收目标"]
    for col, header in enumerate(milestone_headers, 1):
        milestones.cell(4, col, header)
    style_header(milestones[4])
    for row_index, values in enumerate(MILESTONES, 5):
        for col_index, value in enumerate(values, 1):
            cell = milestones.cell(row_index, col_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="thin", color=GREY_200))
            if isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
        milestones.row_dimensions[row_index].height = 62
    milestones.freeze_panes = "A5"
    for col, width in zip("ABCDEF", [12, 24, 16, 14, 62, 48]):
        milestones.column_dimensions[col].width = width
    milestone_table = Table(displayName="DemoMilestones", ref=f"A4:F{4 + len(MILESTONES)}")
    milestone_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    milestones.add_table(milestone_table)

    # 字段说明
    fields = wb.create_sheet("字段说明")
    style_title(fields, "路线图字段说明", "字段以跨部门可理解、可筛选、可验收为原则。", 4)
    for col, header in enumerate(["字段", "建议类型", "填写说明", "要求"], 1):
        fields.cell(4, col, header)
    style_header(fields[4])
    for row_index, values in enumerate(FIELD_GUIDE, 5):
        for col_index, value in enumerate(values, 1):
            cell = fields.cell(row_index, col_index, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color=GREY_200))
        fields.row_dimensions[row_index].height = 44
    fields.freeze_panes = "A5"
    for col, width in zip("ABCD", [22, 18, 80, 16]):
        fields.column_dimensions[col].width = width
    field_table = Table(displayName="RoadmapFieldGuide", ref=f"A4:D{4 + len(FIELD_GUIDE)}")
    field_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    fields.add_table(field_table)

    # 保存并回读校验
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    check = load_workbook(OUTPUT, data_only=False)
    assert check.sheetnames == ["路线图仪表盘", "全部功能路线图", "枚举字典", "里程碑", "字段说明"]
    assert check["全部功能路线图"].max_row == 4 + len(rows)
    assert check["全部功能路线图"].max_column == len(HEADERS)
    check.close()


if __name__ == "__main__":
    build_workbook()
    print(f"Generated: {OUTPUT}")
    print(f"Roadmap items: {len(rows)}")
